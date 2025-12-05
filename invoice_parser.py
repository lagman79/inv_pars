import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import os
import json
import threading
import time
import re
import shutil  # <--- Επανήλθε για την μετακίνηση αρχείων
import pandas as pd
import google.generativeai as genai
from glob import glob
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill

# --- ΡΥΘΜΙΣΕΙΣ ---
SETTINGS_FILE = "settings.json"

RULES_GREEN = """
ΚΑΝΟΝΕΣ ΚΑΤΗΓΟΡΙΟΠΟΙΗΣΗΣ (category_code):
- 1: Κτιριακές και λοιπές εγκαταστάσεις.
- 2.1: Παραγωγικός & Μηχανολογικός Εξοπλισμός.
- 2.2: Λοιπός εξοπλισμός (συστήματα ασφαλείας), εξοπλισμός γραφείου.
- 3: Εξοπλισμός για βελτίωση Ενεργειακής Απόδοσης.
- 4.1: Πιστοποίηση/συμμόρφωση προϊόντων.
- 4.2: Πιστοποίηση υπηρεσιών & διαδικασιών.
- 4.3: Πνευματική ιδιοκτησία - Ευρεσιτεχνίες.
- 5: Υπηρεσίες Σχεδιασμού Συσκευασίας - Branding.
- 6: Δαπάνες Προβολής και Εξωστρέφειας.
- 7: Συμμετοχή σε εμπορικές εκθέσεις.
- 8.1: Τεχνικές μελέτες.
- 8.2: Συμβουλευτική Υποστήριξη.
- 9: Ηλεκτρικά Μεταφορικά Μέσα.
- 10: Δαπάνες Προσωπικού.
"""

RULES_DIGITAL = """
ΚΑΝΟΝΕΣ ΚΑΤΗΓΟΡΙΟΠΟΙΗΣΗΣ (category_code):
- 1: Ψηφιακός εξοπλισμός γραφείου (PC, Laptop, Printers, Servers).
- 2: Αναβάθμιση υποδομών internet (Cabling, Wi-Fi).
- 3: Εφαρμογές γραφείου/ασφάλειας/αποθήκευσης (Office, Antivirus, Cloud).
- 4: Εφαρμογές βελτιστοποίησης (ERP, CRM, WMS, E-shop).
- 5: Κατασκευή ιστοσελίδας, eshop.
- 6: Συμβουλευτική υποστήριξη.
- 7: Τεχνικός σύμβουλος.
- 8: Λογισμικό ως υπηρεσία (SaaS).
"""

class DataProcessor:
    
    @staticmethod
    def sanitize_filename(name):
        """Καθαρίζει το όνομα αρχείου από απαγορευμένους χαρακτήρες (Για Μετονομασία)"""
        if not name: return "Unknown"
        clean = re.sub(r'[\\/*?:"<>|]', '_', str(name))
        return clean.strip()

    @staticmethod
    def validate_mark_code(code):
        if not code: return None
        clean_text = str(code).replace(" ", "").replace("-", "")
        match = re.search(r'40\d{13,}', clean_text)
        return match.group(0) if match else None

    @staticmethod
    def normalize_type(t):
        if not t: return "ΤΙΜ"
        t = t.strip().upper()
        valid = ["ΤΙΜ", "ΤΠΥ", "ΤΔΑ", "ΠΙΣΤΩΤΙΚΟ", "ΠΑΡΑΓΓΕΛΙΑ"]
        return t if t in valid else "ΤΙΜ"

    @staticmethod
    def format_currency(val):
        if not val: return ""
        try:
            if any(c.isalpha() for c in str(val)): return str(val)
            return "{:,.2f}".format(float(val)).replace(",", "X").replace(".", ",").replace("X", ".") + " €"
        except: return str(val)

    @staticmethod
    def format_date(val):
        """Μετατροπή ISO (YYYY-MM-DD) σε DD/MM/YYYY"""
        if not val: return ""
        try:
            dt = pd.to_datetime(val, errors='coerce')
            if pd.notnull(dt):
                return dt.strftime('%d/%m/%Y')
            return str(val)
        except:
            return str(val)

    @staticmethod
    def fix_description_lines(data):
        desc = data.get('description')
        if isinstance(desc, list):
            items = []
            for item in desc:
                if isinstance(item, dict): items.append(str(item.get('description', '')))
                else: items.append(str(item))
            data['description'] = " | ".join(items)

        keys_to_check = ['lines', 'items', 'products']
        found_extra_desc = []
        for key in keys_to_check:
            if key in data and isinstance(data[key], list):
                for item in data[key]:
                    if isinstance(item, dict):
                        d = item.get('description')
                        if d: found_extra_desc.append(str(d))
                    else: found_extra_desc.append(str(item))
                del data[key]
        
        if (not data.get('description') or data.get('description') == "") and found_extra_desc:
            data['description'] = " | ".join(found_extra_desc)

        return data

    @staticmethod
    def analyze_file(file_path, api_key, mode, full_extract):
        genai.configure(api_key=api_key, transport='rest')
        sample_file = genai.upload_file(path=file_path, display_name="Invoice")
        
        timeout = 60 
        start_time = time.time()
        while sample_file.state.name == "PROCESSING":
            if time.time() - start_time > timeout: raise TimeoutError("Timeout Google AI.")
            time.sleep(1)
            sample_file = genai.get_file(sample_file.name)
        
        if sample_file.state.name == "FAILED": raise ValueError("Το Google AI δεν μπόρεσε να διαβάσει το αρχείο.")

        model = genai.GenerativeModel("models/gemini-2.5-flash", generation_config={"response_mime_type": "application/json"})

        cat_rules = ""
        if mode == "Πράσινη Παραγωγική Επένδυση ΜμΕ": cat_rules = RULES_GREEN
        elif mode == "Βασικός Ψηφιακός Μετασχηματισμός ΜμΕ": cat_rules = RULES_DIGITAL
        else: cat_rules = "- 1: Hardware\n- 3: Software\n- 8: SaaS\n- 6: Services"

        extra_instruction = ""
        if full_extract:
            extra_instruction = """
            ΕΠΙΠΛΕΟΝ (FULL EXTRACT):
            Ψάξε για ΟΠΟΙΟΔΗΠΟΤΕ άλλο πεδίο υπάρχει.
            Βάλτα όλα σε ένα αντικείμενο 'dynamic_fields' με κλειδιά ΑΥΣΤΗΡΑ ΣΤΑ ΕΛΛΗΝΙΚΑ.
            """

        prompt = f"""
        Είσαι ειδικός λογιστής. Ανάλυσε το παραστατικό και δώσε JSON.
        
        {cat_rules}

        ΟΔΗΓΙΕΣ:
        1. date: Επίστρεψε την ημερομηνία ΑΥΣΤΗΡΑ σε μορφή YYYY-MM-DD (π.χ. 2025-05-03).
        2. serial_number (SERIAL): Ψάξε για "s/n", "serial" ή ορφανά αλφαριθμητικά κάτω από την περιγραφή.
        3. net_value: Η Φορολογητέα Αξία (μετά τις εκπτώσεις).
        4. extra_charges: Ψάξε για "Κρατήσεις", "Παρακράτηση Φόρου", "Χαρτόσημο", "Εισφορά", "Λοιπά Έξοδα", "Περιβαλλοντικό τέλος". Αν βρεις, γράψε περιγραφή και ποσό (π.χ. "Παρακράτηση 20%: 40€"). Αν είναι πολλά, χώρισε με "|". Αν δεν υπάρχουν, null.
        5. mark_code: Ψάξε αριθμό που ξεκινάει με "40" (Regex search).
        6. description: Ενιαίο string με "|".
        7. type: "ΤΠΥ", "ΤΔΑ", "ΠΙΣΤΩΤΙΚΟ", "ΤΙΜ", "ΠΑΡΑΓΓΕΛΙΑ".
        
        {extra_instruction}

        ΠΕΔΙΑ JSON:
        - date (Format: YYYY-MM-DD), supplier_name, supplier_vat, invoice_number, mark_code
        - description (String), net_value, vat_value, total_amount, type
        - related_document, notes, loading_place, destination_place
        - category_code, serial_number
        - extra_charges (String ή null)
        {"- dynamic_fields (object)" if full_extract else ""}
        """

        try:
            response = model.generate_content([sample_file, prompt])
            genai.delete_file(sample_file.name)
            data = json.loads(response.text)

            if isinstance(data, list):
                if len(data) > 0: data = data[0]
                else: return None 

            data = DataProcessor.fix_description_lines(data)
            data['mark_code'] = DataProcessor.validate_mark_code(data.get('mark_code'))
            data['type'] = DataProcessor.normalize_type(data.get('type'))
            
            return data
        except Exception as e:
            try: genai.delete_file(sample_file.name)
            except: pass
            raise e

    @staticmethod
    def get_friendly_error(e):
        msg = str(e)
        if "list" in msg and "get" in msg: return "Μη αναμενόμενη μορφή δεδομένων (AI Error)."
        if "JSON" in msg: return "Το AI δεν επέστρεψε έγκυρο JSON."
        if "429" in msg: return "Όριο αιτημάτων Google (Too Many Requests)."
        if "403" in msg or "API key" in msg: return "Πρόβλημα με το API Key."
        if "Timeout" in msg: return "Το Google AI αργεί να απαντήσει."
        return f"Άγνωστο Σφάλμα: {msg}"

class InvoiceApp:
    def __init__(self, root):
        self.root = root
        self.root.title("AI Invoice Extractor")
        self.root.geometry("800x900")
        
        self.input_folder = tk.StringVar()
        self.output_file = tk.StringVar()
        self.api_key = tk.StringVar()
        self.extract_all = tk.BooleanVar()
        self.category_mode = tk.StringVar(value="None")
        
        # --- Variables for Organize (Restored) ---
        self.do_rename = tk.BooleanVar()
        self.do_organize = tk.BooleanVar()
        self.organize_target_folder = tk.StringVar()
        self.organize_action = tk.StringVar(value="copy")

        self.is_running = False
        
        self.load_settings()
        self.create_widgets()
        
    def create_widgets(self):
        frame_api = tk.LabelFrame(self.root, text="🔐 Ρυθμίσεις Ασφαλείας", padx=10, pady=10)
        frame_api.pack(fill="x", padx=10, pady=5)
        tk.Label(frame_api, text="Gemini API Key:").pack(side="left")
        self.entry_api = tk.Entry(frame_api, textvariable=self.api_key, show="*", width=50)
        self.entry_api.pack(side="left", padx=5)
        self.add_context_menu(self.entry_api)

        frame_files = tk.LabelFrame(self.root, text="📂 Διαχείριση Αρχείων", padx=10, pady=10)
        frame_files.pack(fill="x", padx=10, pady=5)
        
        tk.Button(frame_files, text="Επιλογή Φακέλου", command=self.select_input, width=20).grid(row=0, column=0, pady=2)
        tk.Entry(frame_files, textvariable=self.input_folder, width=55, state="readonly").grid(row=0, column=1, padx=5)
        tk.Label(frame_files, text="(PDF, JPG, PNG, WEBP)", font=("Arial", 8, "italic"), fg="gray").grid(row=1, column=1, sticky="w", padx=5)
        
        tk.Button(frame_files, text="Αποθήκευση Excel", command=self.select_output, width=20).grid(row=2, column=0, pady=2)
        tk.Entry(frame_files, textvariable=self.output_file, width=55, state="readonly").grid(row=2, column=1, padx=5)

        frame_opts = tk.LabelFrame(self.root, text="⚙️ Παράμετροι", padx=10, pady=10)
        frame_opts.pack(fill="x", padx=10, pady=5)

        tk.Label(frame_opts, text="Πρόγραμμα:").grid(row=0, column=0, sticky="w")
        options = ["Χωρίς Κατηγοριοποίηση", "Πράσινη Παραγωγική Επένδυση ΜμΕ", "Βασικός Ψηφιακός Μετασχηματισμός ΜμΕ"]
        self.combo_cat = ttk.Combobox(frame_opts, textvariable=self.category_mode, values=options, width=45, state="readonly")
        self.combo_cat.current(0)
        self.combo_cat.grid(row=0, column=1, padx=5, sticky="w")

        tk.Checkbutton(frame_opts, text="Ενεργοποίηση Full Extract (Δυναμικά Πεδία)", variable=self.extract_all).grid(row=1, column=0, columnspan=2, sticky="w", pady=5)

        # --- ORGANIZE FRAME (RESTORED) ---
        frame_org = tk.LabelFrame(self.root, text="🗂️ Οργάνωση & Μετονομασία", padx=10, pady=10, bg="#f9f9f9")
        frame_org.pack(fill="x", padx=10, pady=5)

        tk.Checkbutton(frame_org, text="Αυτόματη Μετονομασία (Προμηθευτής_Αρ.Παραστατικού)", variable=self.do_rename, bg="#f9f9f9").grid(row=0, column=0, columnspan=2, sticky="w")

        tk.Checkbutton(frame_org, text="Ταξινόμηση σε Φακέλους (ανά Προμηθευτή)", variable=self.do_organize, command=self.toggle_organize_ui, bg="#f9f9f9").grid(row=1, column=0, columnspan=2, sticky="w")
        
        self.btn_target_folder = tk.Button(frame_org, text="Φάκελος Προορισμού", command=self.select_organize_folder, state="disabled")
        self.btn_target_folder.grid(row=2, column=0, padx=5, pady=2)
        
        self.entry_target_folder = tk.Entry(frame_org, textvariable=self.organize_target_folder, width=40, state="disabled")
        self.entry_target_folder.grid(row=2, column=1, padx=5, pady=2)

        self.frame_radio = tk.Frame(frame_org, bg="#f9f9f9")
        self.frame_radio.grid(row=3, column=0, columnspan=2, sticky="w", padx=5)
        self.rb_copy = tk.Radiobutton(self.frame_radio, text="Αντιγραφή (Copy)", variable=self.organize_action, value="copy", bg="#f9f9f9", state="disabled")
        self.rb_copy.pack(side="left")
        self.rb_move = tk.Radiobutton(self.frame_radio, text="Μετακίνηση (Cut)", variable=self.organize_action, value="move", bg="#f9f9f9", state="disabled")
        self.rb_move.pack(side="left")

        self.lbl_progress_text = tk.Label(self.root, text="Αναμονή εκκίνησης...", font=("Arial", 9))
        self.lbl_progress_text.pack(anchor="w", padx=15)

        self.progress = ttk.Progressbar(self.root, orient="horizontal", length=100, mode="determinate")
        self.progress.pack(fill="x", padx=15, pady=5)
        
        self.log_text = tk.Text(self.root, height=10, state="disabled", bg="#1e1e1e", fg="#00ff00", font=("Consolas", 9))
        self.log_text.pack(fill="both", expand=True, padx=10, pady=5)

        self.btn_start = tk.Button(self.root, text="🚀 ΕΚΚΙΝΗΣΗ", command=self.start_thread, bg="#2ecc71", fg="white", font=("Arial", 12, "bold"))
        self.btn_start.pack(fill="x", padx=10, pady=5)

        frame_actions = tk.Frame(self.root)
        frame_actions.pack(fill="x", padx=10, pady=5)
        tk.Button(frame_actions, text="🔄 Νέα Εργασία", command=self.reset_app, width=15).pack(side="left", padx=5)
        tk.Button(frame_actions, text="❌ Έξοδος", command=self.close_app, width=15, fg="red").pack(side="right", padx=5)

    def toggle_organize_ui(self):
        state = "normal" if self.do_organize.get() else "disabled"
        self.btn_target_folder.config(state=state)
        self.rb_copy.config(state=state)
        self.rb_move.config(state=state)

    def add_context_menu(self, widget):
        menu = tk.Menu(self.root, tearoff=0)
        menu.add_command(label="Επικόλληση", command=lambda: widget.event_generate("<<Paste>>"))
        widget.bind("<Button-3>", lambda e: menu.post(e.x_root, e.y_root))

    def log(self, msg, color=None):
        self.log_text.config(state="normal")
        tag_name = "normal"
        if color:
            tag_name = color
            self.log_text.tag_config(color, foreground=color)
        self.log_text.insert("end", f"[{datetime.now().strftime('%H:%M:%S')}] {msg}\n", tag_name)
        self.log_text.see("end")
        self.log_text.config(state="disabled")

    def load_settings(self):
        if os.path.exists(SETTINGS_FILE):
            try:
                with open(SETTINGS_FILE, "r") as f:
                    data = json.load(f)
                    self.api_key.set(data.get("api_key", ""))
                    self.input_folder.set(data.get("input_folder", ""))
                    self.output_file.set(data.get("output_file", ""))
                    self.organize_target_folder.set(data.get("organize_target_folder", ""))
            except: pass

    def save_settings(self):
        data = {
            "api_key": self.api_key.get(),
            "input_folder": self.input_folder.get(),
            "output_file": self.output_file.get(),
            "organize_target_folder": self.organize_target_folder.get()
        }
        with open(SETTINGS_FILE, "w") as f:
            json.dump(data, f)

    def select_input(self):
        f = filedialog.askdirectory()
        if f: self.input_folder.set(f)

    def select_output(self):
        f = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
        if f: self.output_file.set(f)

    def select_organize_folder(self):
        f = filedialog.askdirectory()
        if f: self.organize_target_folder.set(f)
    
    def reset_app(self):
        self.input_folder.set("")
        self.output_file.set("")
        self.progress["value"] = 0
        self.lbl_progress_text.config(text="Αναμονή εκκίνησης...")
        self.log_text.config(state="normal")
        self.log_text.delete(1.0, "end")
        self.log_text.config(state="disabled")
        self.btn_start.config(state="normal")
        self.log("🔄 Έτοιμο για νέα εργασία.")

    def close_app(self):
        if messagebox.askyesno("Έξοδος", "Θέλετε σίγουρα να κλείσετε την εφαρμογή;"):
            self.root.destroy()

    def start_thread(self):
        if not self.api_key.get() or not self.input_folder.get():
            messagebox.showwarning("Προσοχή", "Λείπουν στοιχεία (API Key ή Φάκελος)!")
            return
        
        if self.do_organize.get() and not self.organize_target_folder.get():
            messagebox.showwarning("Προσοχή", "Επιλέξτε Φάκελο Προορισμού για την ταξινόμηση!")
            return

        self.save_settings()
        self.is_running = True
        self.btn_start.config(state="disabled", text="⏳ ΣΕ ΕΞΕΛΙΞΗ...")
        self.log_text.config(state="normal"); self.log_text.delete(1.0, "end"); self.log_text.config(state="disabled")
        
        threading.Thread(target=self.run_process, daemon=True).start()

    def run_process(self):
        try:
            input_dir = self.input_folder.get()
            output_path = self.output_file.get()
            if not output_path: output_path = os.path.join(input_dir, "results.xlsx")
            
            supported_extensions = ["*.pdf", "*.jpg", "*.jpeg", "*.png", "*.webp"]
            all_files = []
            for ext in supported_extensions:
                all_files.extend(glob(os.path.join(input_dir, ext)))
            all_files = sorted(all_files)

            if not all_files:
                self.log("❌ Δεν βρέθηκαν υποστηριζόμενα αρχεία.", "red")
                return

            total_files = len(all_files)
            self.progress["maximum"] = total_files
            all_data = []
            success_count = 0
            fail_count = 0

            for i, f in enumerate(all_files, 1):
                if not self.is_running: break
                
                filename = os.path.basename(f)
                self.lbl_progress_text.config(text=f"Επεξεργασία: {i} από {total_files} ({filename})")
                self.log(f"[{i}/{total_files}] Ανάλυση: {filename}")
                
                try:
                    data = DataProcessor.analyze_file(
                        f, 
                        self.api_key.get().strip(), 
                        self.category_mode.get(), 
                        self.extract_all.get()
                    )
                    if data:
                        data['filename'] = filename
                        data['original_path'] = f # Keep for rename/move
                        data['processing_status'] = 'OK'
                        all_data.append(data)
                        success_count += 1
                        self.log("   ✅ Επιτυχία", "#00ff00")
                    else:
                        raise ValueError("Κενή απάντηση από AI")
                except Exception as e:
                    fail_count += 1
                    friendly_error = DataProcessor.get_friendly_error(e)
                    self.log(f"   ❌ Σφάλμα: {friendly_error}", "#ff5555")
                    all_data.append({
                        'filename': filename,
                        'processing_status': 'FAILED',
                        'notes': friendly_error
                    })
                
                self.progress["value"] = i
                time.sleep(2)

            # --- POST PROCESSING (Rename & Organize) ---
            if self.do_rename.get() or self.do_organize.get():
                self.log("\n⚙️ Έναρξη Οργάνωσης Αρχείων...", "cyan")
                self.post_process_files(all_data)

            self.lbl_progress_text.config(text=f"Ολοκληρώθηκε: {total_files} αρχεία.")
            self.log("\n" + "="*30)
            self.log(f"📊 ΑΠΟΤΕΛΕΣΜΑΤΑ:")
            self.log(f"✅ Επιτυχημένα: {success_count}", "#00ff00")
            self.log(f"❌ Αποτυχημένα: {fail_count}", "#ff5555")
            self.log("="*30 + "\n")

            if all_data:
                self.generate_excel(all_data, output_path)
            else:
                self.log("⚠️ Δεν δημιουργήθηκε αρχείο (Κενά δεδομένα).", "yellow")

        except Exception as e:
            self.log(f"ΚΡΙΣΙΜΟ ΣΦΑΛΜΑ: {e}", "red")
            messagebox.showerror("Error", str(e))
        finally:
            self.is_running = False
            self.btn_start.config(state="normal", text="🚀 ΕΚΚΙΝΗΣΗ")

    def post_process_files(self, all_data):
        target_dir = self.organize_target_folder.get()
        action = self.organize_action.get()

        for item in all_data:
            if item.get('processing_status') != 'OK': continue
            
            current_path = item.get('original_path')
            if not current_path or not os.path.exists(current_path): continue

            supplier = item.get('supplier_name', 'Unknown')
            invoice_num = item.get('invoice_number', '000')
            ext = os.path.splitext(current_path)[1]

            # 1. RENAME
            if self.do_rename.get():
                safe_supplier = DataProcessor.sanitize_filename(supplier)
                safe_invoice = DataProcessor.sanitize_filename(invoice_num)
                new_name = f"{safe_supplier}_{safe_invoice}{ext}"
                
                dir_name = os.path.dirname(current_path)
                new_path = os.path.join(dir_name, new_name)
                
                try:
                    if os.path.exists(new_path) and new_path != current_path:
                        new_name = f"{safe_supplier}_{safe_invoice}_{int(time.time())}{ext}"
                        new_path = os.path.join(dir_name, new_name)

                    os.rename(current_path, new_path)
                    current_path = new_path
                    item['filename'] = new_name
                    self.log(f"   ✏️ Μετονομασία: {new_name}")
                except Exception as e:
                    self.log(f"   ⚠️ Rename Error: {e}", "yellow")

            # 2. ORGANIZE
            if self.do_organize.get():
                safe_supplier = DataProcessor.sanitize_filename(supplier)
                dest_folder = os.path.join(target_dir, safe_supplier)
                
                try:
                    os.makedirs(dest_folder, exist_ok=True)
                    dest_path = os.path.join(dest_folder, os.path.basename(current_path))
                    
                    if action == 'move':
                        shutil.move(current_path, dest_path)
                        self.log(f"   🚚 Μετακίνηση -> {safe_supplier}")
                    else:
                        shutil.copy2(current_path, dest_path)
                        self.log(f"   📋 Αντιγραφή -> {safe_supplier}")
                except Exception as e:
                    self.log(f"   ⚠️ Org Error: {e}", "yellow")

    def generate_excel(self, all_data, path):
        df = pd.DataFrame(all_data)
        
        for col in ['net_value', 'vat_value', 'total_amount']:
            if col in df.columns:
                df[col] = df[col].apply(DataProcessor.format_currency)

        mapping = {
            'supplier_name': 'ΠΡΟΜΗΘΕΥΤΗΣ', 'supplier_vat': 'ΑΦΜ ΠΡΟΜΗΘΕΥΤΗ',
            'type': 'ΕΙΔΟΣ ΠΑΡΑΣΤ.', 'invoice_number': 'ΑΡ. ΠΑΡΑΣΤ.',
            'date': 'ΗΜ/ΝΙΑ', 'description': 'ΠΕΡΙΓΡΑΦΗ',
            'net_value': 'ΚΑΘΑΡΗ ΑΞΙΑ', 'vat_value': 'ΦΠΑ',
            'total_amount': 'ΤΕΛΙΚΗ ΑΞΙΑ', 'mark_code': 'ΜΑΡΚ',
            'filename': 'ΟΝΟΜΑ ΑΡΧΕΙΟΥ', 'serial_number': 'SERIAL',
            'category_code': 'Κ.Δ.', 'related_document': 'ΣΧΕΤ. ΠΑΡΑΣΤ.',
            'notes': 'ΠΑΡΑΤΗΡΗΣΕΙΣ', 'loading_place': 'ΦΟΡΤΩΣΗ',
            'destination_place': 'ΠΡΟΟΡΙΣΜΟΣ',
            'extra_charges': 'ΚΡΑΤΗΣΕΙΣ / ΕΞΟΔΑ' # <--- Included
        }
        
        dynamic_cols = []
        if self.extract_all.get() and 'dynamic_fields' in df.columns:
            def get_series(x): return pd.Series(x) if isinstance(x, dict) else pd.Series()
            dynamic_df = df['dynamic_fields'].apply(get_series)
            df = pd.concat([df.drop(['dynamic_fields'], axis=1), dynamic_df], axis=1)
            dynamic_cols = list(dynamic_df.columns)

        df.rename(columns=mapping, inplace=True)

        # ISO Date Fix applied here
        if 'ΗΜ/ΝΙΑ' in df.columns:
            df['ΗΜ/ΝΙΑ'] = df['ΗΜ/ΝΙΑ'].apply(DataProcessor.format_date)

        left_cols = [
            'ΠΡΟΜΗΘΕΥΤΗΣ', 'ΑΦΜ ΠΡΟΜΗΘΕΥΤΗ', 'ΕΙΔΟΣ ΠΑΡΑΣΤ.', 'ΑΡ. ΠΑΡΑΣΤ.',
            'ΗΜ/ΝΙΑ', 'ΠΕΡΙΓΡΑΦΗ', 'ΚΑΘΑΡΗ ΑΞΙΑ', 'ΦΠΑ', 'ΤΕΛΙΚΗ ΑΞΙΑ',
            'ΜΑΡΚ', 'ΟΝΟΜΑ ΑΡΧΕΙΟΥ', 'ΣΧΕΤ. ΠΑΡΑΣΤ.', 'ΠΑΡΑΤΗΡΗΣΕΙΣ', 
            'ΦΟΡΤΩΣΗ', 'ΠΡΟΟΡΙΣΜΟΣ', 'Κ.Δ.', 'SERIAL' 
        ]
        
        right_cols = ['ΚΡΑΤΗΣΕΙΣ / ΕΞΟΔΑ']

        existing_left = [c for c in left_cols if c in df.columns]
        existing_right = [c for c in right_cols if c in df.columns]
        existing_dynamic = [c for c in df.columns if c not in left_cols and c not in right_cols and c != 'processing_status']
        
        final_order = existing_left + existing_dynamic + existing_right
        
        df = df[final_order]
        df = df.fillna("") 
        
        try:
            df.to_excel(path, index=False)
            self.apply_excel_styles(path, all_data)
            self.log(f"🎉 Το Excel αποθηκεύτηκε: {path}", "#00ff00")
            
            if messagebox.askyesno("Ολοκληρώθηκε", f"Η διαδικασία τελείωσε!\nΑρχείο: {path}\n\nΘέλετε να ανοίξετε το Excel τώρα;"):
                os.startfile(path)
        except PermissionError:
            messagebox.showerror("Σφάλμα", "Κλείσε το αρχείο Excel! Είναι ανοιχτό.")

    def apply_excel_styles(self, path, all_data):
        try:
            wb = load_workbook(path)
            ws = wb.active
            red_fill = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")
            for i, item in enumerate(all_data):
                if item.get('processing_status') == 'FAILED':
                    row_idx = i + 2 
                    for cell in ws[row_idx]: cell.fill = red_fill
            wb.save(path)
        except Exception as e:
            self.log(f"Warning: Δεν μπόρεσα να χρωματίσω το Excel ({e})", "yellow")

if __name__ == "__main__":
    root = tk.Tk()
    app = InvoiceApp(root)
    root.mainloop()